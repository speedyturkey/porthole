import os
import unittest
from porthole import WorkbookEditor
import openpyxl


class TestWorkbookEditor(unittest.TestCase):
    """
    report.workbook.worksheets()[0].name
    sheetname_count - count of sheets
    sheetnames - dictionary where keys are sheetnames
    fileclosed - 1 if closed, 0 otherwise
    """

    def setUp(self):
        self.test_workbook = openpyxl.Workbook()
        self.test_sheet = self.test_workbook.active
        self.test_sheet.title = 'test_sheet_1'
        self.test_sheet['A1'] = 3.14
        self.test_fields = ['a', 'b', 'c', 'd', 'e', 'f']
        self.test_data = [
            ('the', 'grand', 'old', 'duke', 'of', 'york'), ('had', 'ten', 'thousand', 'men', 'he', 'marched')
        ]
        self.test_filename = 'testfile.xlsx'
        self.test_workbook.save(filename=self.test_filename)

    def tearDown(self):
        os.remove(self.test_filename)

    def test_initialize_workbook_editor(self):
        # with correct filename
        correct_wb = WorkbookEditor(workbook_filename = self.test_filename )
        self.assertIsInstance(correct_wb, WorkbookEditor)

        # with incorrect filename
        try:
            incorrect_wb = WorkbookEditor(workbook_filename='this_file_does_not_exist.xlsx')
            raise Exception('This method should have raised FileNotFoundError.')
        except FileNotFoundError:
            pass
        except Exception as e:
            raise Exception('This method should have raised FileNotFoundError.  Instead it raised: {}'.format(e))

    def test_replace_sheet_contents(self):
        sheet_name = self.test_sheet.title
        wb = WorkbookEditor(workbook_filename=self.test_filename)
        wb.replace_sheet_contents(sheet_name=sheet_name, data_rows=self.test_data, headers=self.test_fields)
        values = wb.get_all_values(sheet_name)
        self.assertEqual(values, [tuple(self.test_fields)] + self.test_data)

    def test_save_workbook(self):
        wb = WorkbookEditor(workbook_filename=self.test_filename)
        sheet_name = self.test_sheet.title

        # with save_as defined
        save_as = 'foo.xlsx'
        wb.save_workbook(save_as=save_as)
        wb2 = WorkbookEditor(workbook_filename=save_as)
        values = wb.get_all_values(sheet_name)
        self.assertEqual(values[0][0], 3.14)

        # with save_as undefined
        os.remove(self.test_filename)
        wb.save_workbook()
        wb2 = WorkbookEditor(workbook_filename=self.test_filename)
        values = wb2.get_all_values(sheet_name)
        self.assertEqual(values[0][0], 3.14)


